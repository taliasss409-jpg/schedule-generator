import io
import re
from collections import defaultdict

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


# ============================ הגדרות קבועות ============================

# ערוץ הוא תפקיד שקיים רק בצהריים, ולא משמרת נפרדת.
SHIFTS = ['בוקר', 'צהריים', 'ערב', 'לילה']

OUTPUT_COLUMNS = [
    'יום', 'תאריך', 'משמרת', 'מנמ"ש', 'ווטסאפ', 'צנזור 1', 'צנזור 2',
    'מנהל ידע', "צ'קינג 11", "צ'קינג 12", "צ'קינג 13", "צ'קינג 14",
    "צ'קינג 24I", 'רח"ל', 'דסקאי 1', 'דסקאי 2',
    'ערוץ 11', 'ערוץ 12', 'ערוץ 13', 'ערוץ 14', 'מדור', 'חופש', 'אחרי לילה'
]

MANUAL_ONLY_COLUMNS = {'דסקאי 1', 'דסקאי 2', 'מדור', 'חופש', 'אחרי לילה'}

# סדר מילוי תפקידי הצנזור. כל עובד משובץ פעם אחת בלבד.
CENSOR_ASSIGNMENT_ORDER = [
    'ווטסאפ',
    'צנזור 1',
    'צנזור 2',
    "צ'קינג 11",
    "צ'קינג 12",
    "צ'קינג 13",
    "צ'קינג 14",
    "צ'קינג 24I",
    'רח"ל',
]

# רק אחרי שכל תפקידי הצנזור מולאו, עודפים מתווספים לכאן.
CENSOR_OVERFLOW_COLUMN = 'צנזור 2'

# ערוץ קיים רק בצהריים ורק לצנזור שביקש במפורש "ערוץ".
CHANNEL_COLUMNS = ['ערוץ 11', 'ערוץ 12', 'ערוץ 13', 'ערוץ 14']

_CENSOR_COLUMNS = CENSOR_ASSIGNMENT_ORDER + CHANNEL_COLUMNS

ROLE_TO_COLUMNS = {
    'צנזור': _CENSOR_COLUMNS,
    'סדיר (קבע/אע"צ)': _CENSOR_COLUMNS,
    'מנהל משמרת': ['מנמ"ש'],
    'מנהל ידע': ['מנהל ידע', 'צנזור 1', 'צנזור 2'],
    'מילואים': _CENSOR_COLUMNS,
}

# אלה העובדים שמחולקים פעם אחת בלבד בין כל תפקידי הצנזור.
CENSOR_ROLES = {'צנזור', 'סדיר (קבע/אע"צ)', 'מילואים'}

SHIFT_ALIASES = {
    'אחרי משמרת לילה': 'לילה',
}

# "ערוץ" אינו ב-SHIFTS בכוונה: הוא מטופל בנפרד כאפשרות צהריים.
IGNORE_TOKENS = {'חופש', 'מרוכז', 'יום חופש מרוכז'}


def get_role(service_type, main_role):
    """קובעת את התפקיד בפועל: תפקיד עיקרי אם קיים, אחרת סוג שירות."""
    for value in (main_role, service_type):
        if value is not None:
            text = str(value).strip()
            if text and text.lower() != 'nan':
                return text
    return None


def parse_availability(cell_value):
    """הופכת טקסט חופשי כמו 'בוקר, אחרי משמרת לילה' לרשימת משמרות מנורמלת (set)."""
    if cell_value is None:
        return set()
    text = str(cell_value).strip()
    if not text or text.lower() == 'nan':
        return set()
    shifts = set()
    for raw in text.split(','):
        token = raw.strip()
        if not token or token in IGNORE_TOKENS:
            continue
        token = SHIFT_ALIASES.get(token, token)
        if token in SHIFTS or token == 'ערוץ':
            shifts.add(token)
        # טוקן לא מזוהה - מתעלמים בשקט (אפשר להרחיב טיפול בהמשך)
    return shifts


def extract_days_from_requests(columns):
    """
    מזהה עמודות זמינות בצורה גמישה.

    תומך למשל ב:
    זמינות א' 23.08
    זמינות א׳ 23.08
    זמינות א 23.08
    זמינות א' - 23.08
    זמינות א׳-23.08
    """
    days = []

    # מחפשים "זמינות", אחריה אות יום, ואז תאריך.
    # מאפשרים סוגים שונים של גרש, מקפים ורווחים.
    pattern = re.compile(
        r"זמינות\s*([א-ת])(?:['׳״\"])?\s*[-–—]?\s*"
        r"(\d{1,2}[./-]\d{1,2}(?:[./-]\d{2,4})?)"
    )

    for col in columns:
        text = str(col).strip()
        m = pattern.search(text)

        if m:
            days.append((col, m.group(1), m.group(2)))

    # אם לא נמצא כלום, מציגים את הכותרות כדי שיהיה קל לאבחן.
    if not days:
        print("⚠️ לא זוהו עמודות זמינות.")
        print("כותרות שנמצאו בקובץ:")
        for col in columns:
            print(" -", repr(str(col)))

    return days


def build_schedule_workbook(requests_bytes):
    from collections import defaultdict

    df = pd.read_excel(io.BytesIO(requests_bytes))

    day_cols = extract_days_from_requests(df.columns)
    if not day_cols:
        raise ValueError(
            "לא נמצאו עמודות זמינות בקובץ (בפורמט \"זמינות א' 23.08\"). בדקי את כותרות הקובץ."
        )

    # ------------------------------------------------------------
    # בניית אינדקסים מראש: מעבר אחד על העובדים והזמינויות.
    # כך נמנעת סריקה מחדש של כל העובדים עבור כל תא.
    # ------------------------------------------------------------
    employees = []
    regular_index = defaultdict(list)
    censor_index = defaultdict(list)
    channel_index = defaultdict(list)

    for row in df.to_dict(orient='records'):
        first = str(row.get('שם פרטי', '') or '').strip()
        last = str(row.get('שם משפחה', '') or '').strip()
        full_name = f"{first} {last}".strip()

        if not full_name or full_name.lower() == 'nan nan':
            continue

        role = get_role(
            row.get('סוג שירות'),
            row.get('תפקיד עיקרי במפעל')
        )

        availability = {}
        is_censor = role in CENSOR_ROLES

        for col, day_letter, date in day_cols:
            shifts = parse_availability(row.get(col))
            availability[day_letter] = shifts

            # ערוץ: רק צנזור שסימן במפורש "ערוץ".
            if is_censor and 'ערוץ' in shifts:
                channel_index[day_letter].append(full_name)

            for shift in shifts:
                if shift == 'ערוץ':
                    continue

                if is_censor:
                    censor_index[(day_letter, shift)].append(full_name)
                else:
                    for output_column in ROLE_TO_COLUMNS.get(role, []):
                        if output_column not in CHANNEL_COLUMNS:
                            regular_index[(output_column, day_letter, shift)].append(full_name)

        employees.append({
            'name': full_name,
            'role': role,
            'availability': availability
        })

    # ------------------------------------------------------------
    # בניית הוורקבוק
    # ------------------------------------------------------------
    wb = Workbook()
    ws = wb.active
    ws.title = 'סידור עבודה'

    header_font = Font(bold=True)
    header_fill = PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid')
    thin = Side(style='thin', color='BFBFBF')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    header_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
    cell_align = Alignment(horizontal='right', vertical='center', wrap_text=True)

    for c, title in enumerate(OUTPUT_COLUMNS, start=1):
        cell = ws.cell(row=1, column=c, value=title)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = border

    row_idx = 2

    for col, day_letter, date in day_cols:
        first_row_of_day = row_idx

        for shift in SHIFTS:
            row_values = {}

            # ----------------------------------------------------
            # תפקידים רגילים שאינם חלק מחלוקת הצנזורים.
            # ----------------------------------------------------
            for output_column in OUTPUT_COLUMNS:
                if (
                    output_column in MANUAL_ONLY_COLUMNS
                    or output_column in {'יום', 'תאריך', 'משמרת'}
                    or output_column in CENSOR_ASSIGNMENT_ORDER
                    or output_column in CHANNEL_COLUMNS
                ):
                    continue

                candidates = regular_index.get(
                    (output_column, day_letter, shift), []
                )

                if candidates:
                    row_values[output_column] = ', '.join(candidates)

            # ----------------------------------------------------
            # צנזורים: כל אדם מופיע פעם אחת בלבד.
            # ממלאים את כל התיבות לפי הסדר, ורק אז עודפים לצנזור 2.
            # ----------------------------------------------------
            available_censors = censor_index.get(
                (day_letter, shift), []
            )

            for i, output_column in enumerate(CENSOR_ASSIGNMENT_ORDER):
                if i < len(available_censors):
                    row_values[output_column] = available_censors[i]

            if len(available_censors) > len(CENSOR_ASSIGNMENT_ORDER):
                overflow = available_censors[len(CENSOR_ASSIGNMENT_ORDER):]
                existing = row_values.get(CENSOR_OVERFLOW_COLUMN)
                names = ([existing] if existing else []) + overflow
                row_values[CENSOR_OVERFLOW_COLUMN] = ', '.join(names)

            # ----------------------------------------------------
            # ערוץ: רק בשורת צהריים ורק צנזורים שביקשו ערוץ.
            # ----------------------------------------------------
            if shift == 'צהריים':
                channel_candidates = channel_index.get(day_letter, [])

                for i, output_column in enumerate(CHANNEL_COLUMNS):
                    if i < len(channel_candidates):
                        row_values[output_column] = channel_candidates[i]

                # אם יש יותר מועמדים ממספר תיבות הערוץ, שומרים את
                # העודפים בתיבת הערוץ האחרונה כדי לא לאבד זמינות.
                if len(channel_candidates) > len(CHANNEL_COLUMNS):
                    overflow = channel_candidates[len(CHANNEL_COLUMNS):]
                    existing = row_values.get(CHANNEL_COLUMNS[-1])
                    names = ([existing] if existing else []) + overflow
                    row_values[CHANNEL_COLUMNS[-1]] = ', '.join(names)

            # ----------------------------------------------------
            # כתיבת השורה
            # ----------------------------------------------------
            for c, colname in enumerate(OUTPUT_COLUMNS, start=1):
                if colname == 'יום':
                    value = day_letter if row_idx == first_row_of_day else None
                elif colname == 'תאריך':
                    value = date if row_idx == first_row_of_day else None
                elif colname == 'משמרת':
                    value = shift
                else:
                    value = row_values.get(colname)

                cell = ws.cell(row=row_idx, column=c, value=value)
                cell.border = border
                cell.alignment = cell_align

            row_idx += 1

    for c, title in enumerate(OUTPUT_COLUMNS, start=1):
        ws.column_dimensions[get_column_letter(c)].width = max(10, len(title) + 4)

    ws.freeze_panes = 'D2'
    ws.sheet_view.rightToLeft = True

    return wb, employees, day_cols
